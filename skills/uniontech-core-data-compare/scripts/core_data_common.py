#!/usr/bin/env python3
from __future__ import annotations

import json
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path.home()
DOWNLOADS_DIR = ROOT / "Downloads"
OUTPUT_DIR = ROOT / "output"
SKILL_DIR = Path(__file__).resolve().parent.parent
REFERENCES_DIR = SKILL_DIR / "references"


CANONICAL_BLOCK_NAMES = {
    "单个默认插件的被移除率": "单个插件的被移除率",
    "专业版用户-保险箱功能开启率": "专业版用户-保险箱功能开启率（万分之）",
    "专业版用户-全文搜索功能开启率": "专业版用户-全文搜索功能开启率（万分之）",
}


BLOCK_LAYOUTS: dict[str, dict[str, Any]] = {
    "产品分布": {
        "sheet_index": 0,
        "start_col": 1,
        "rows": list(range(4, 9)),
        "col_start": 1,
        "col_end": 2,
        "headers": ["产品", "用户数"],
    },
    "架构分布": {
        "sheet_index": 0,
        "start_col": 3,
        "rows": list(range(4, 9)),
        "col_start": 3,
        "col_end": 4,
        "headers": ["架构", "用户数"],
    },
    "专业版日新增用户数": {
        "sheet_index": 0,
        "start_col": 5,
        "rows": list(range(4, 35)),
        "col_start": 5,
        "col_end": 6,
        "headers": ["日期", "专业版日新增用户数"],
    },
    "专业版累计用户数": {
        "sheet_index": 0,
        "start_col": 7,
        "rows": [4, 11, 18, 25, 32],
        "col_start": 7,
        "col_end": 8,
        "headers": ["日期", "专业版累计用户数"],
    },
    "专业版1070u5累计用户数": {
        "sheet_index": 0,
        "start_col": 9,
        "rows": [4, 11, 18],
        "col_start": 9,
        "col_end": 10,
        "headers": ["日期", "专业版1070u5累计用户数"],
    },
    "专业版V25累计用户数": {
        "sheet_index": 0,
        "start_col": 11,
        "rows": [4, 11, 18, 25, 32],
        "col_start": 11,
        "col_end": 12,
        "headers": ["日期", "专业版V25累计用户数"],
    },
    "专业版版本系列用户分布": {
        "sheet_index": 0,
        "start_col": 13,
        "rows": list(range(4, 28)),
        "col_start": 13,
        "col_end": 14,
        "headers": ["小版本", "用户数"],
    },
    "任务栏模式配置": {
        "sheet_index": 0,
        "start_col": 15,
        "rows": [4, 5],
        "col_start": 15,
        "col_end": 18,
        "headers": ["产品", "版本系列", "高效模式用户数", "时尚模式用户数"],
    },
    "单个插件的被移除率": {
        "sheet_index": 0,
        "start_col": 19,
        "rows": list(range(4, 14)),
        "col_start": 19,
        "col_end": 25,
        "headers": [
            "插件",
            "106x插件被移除率",
            "106x插件被移除数",
            "106x用户数",
            "107x插件被移除率",
            "107x插件被移除数",
            "107x用户数",
        ],
    },
    "专业版用户-保险箱功能开启率（万分之）": {
        "sheet_index": 0,
        "start_col": 26,
        "rows": list(range(4, 14)),
        "col_start": 26,
        "col_end": 29,
        "headers": ["小版本", "保险箱功能开启率", "保险箱功能开启用户数", "总用户数"],
    },
    "专业版用户-全文搜索功能开启率（万分之）": {
        "sheet_index": 0,
        "start_col": 30,
        "rows": list(range(4, 14)),
        "col_start": 30,
        "col_end": 33,
        "headers": ["小版本", "全文搜索功能开启率", "全文搜索功能开启用户数", "总用户数"],
    },
    "专业版应用启动排行Top50-周累计": {
        "sheet_index": 1,
        "start_col": 1,
        "rows": list(range(2, 52)),
        "col_start": 1,
        "col_end": 2,
        "headers": ["应用名称", "启动次数"],
    },
    "教育版应用启动排行Top50-周累计": {
        "sheet_index": 1,
        "start_col": 3,
        "rows": list(range(2, 52)),
        "col_start": 3,
        "col_end": 4,
        "headers": ["应用名称", "启动次数"],
    },
    "专业版检查更新失败率-周累计": {
        "sheet_index": 1,
        "start_col": 5,
        "rows": [4, 5, 6],
        "col_start": 5,
        "col_end": 14,
        "headers": [
            "小版本",
            "LA检查更新失败率",
            "LA检查更新失败数",
            "LA检查更新总数",
            "ARM检查更新失败率",
            "ARM检查更新失败数",
            "ARM检查更新总数",
            "AMD检查更新失败率",
            "AMD检查更新失败数",
            "AMD检查更新总数",
        ],
    },
    "检查更新失败原因分布-周累计": {
        "sheet_index": 1,
        "start_col": 15,
        "rows": list(range(4, 13)),
        "col_start": 15,
        "col_end": 43,
        "headers": [
            "小版本",
            "架构",
            "检查更新失败次数",
            "IndexDownloadFailed_次数",
            "IndexDownloadFailed_占比",
            "JobError::IndexDownloadFailed_次数",
            "JobError::IndexDownloadFailed_占比",
            "platformUnreachable_次数",
            "platformUnreachable_占比",
            "insufficientSpace_次数",
            "insufficientSpace_占比",
            "JobError::ErrorUnknown_次数",
            "JobError::ErrorUnknown_占比",
            "JobError::insufficientSpace_次数",
            "JobError::insufficientSpace_占比",
            "JobError::unmetDependencies_次数",
            "JobError::unmetDependencies_占比",
            "JobError::fetchFailed_次数",
            "JobError::fetchFailed_占比",
            "dpkgError_次数",
            "dpkgError_占比",
            "JobError::invalidSourceList_次数",
            "JobError::invalidSourceList_占比",
            "ErrorUnknown_次数",
            "ErrorUnknown_占比",
            "fetchFailed_次数",
            "fetchFailed_占比",
            "invalidSourceList_次数",
            "invalidSourceList_占比",
        ],
    },
    "smb挂载失败率-周累计": {
        "sheet_index": 1,
        "start_col": 44,
        "rows": [4, 5, 6],
        "col_start": 44,
        "col_end": 53,
        "headers": [
            "小版本",
            "LA smb挂载失败率",
            "LA smb挂载失败次数",
            "LA smb挂载成功+失败次数总和",
            "ARM smb挂载失败率",
            "ARM smb挂载失败次数",
            "ARM smb挂载成功+失败次数总和",
            "AMD smb挂载失败率",
            "AMD smb挂载失败次数",
            "AMD smb挂载成功+失败次数总和",
        ],
    },
    "smb挂载失败原因分布-周累计": {
        "sheet_index": 1,
        "start_col": 54,
        "rows": list(range(4, 13)),
        "col_start": 54,
        "col_end": 60,
        "headers": ["小版本", "架构", "总失败次数", "挂载错误次数", "挂载错误占比", "用户主动取消挂载次数", "用户主动取消挂载占比"],
    },
    "应用商店deb应用新增下载应用排行Top30及下载失败率-周累计": {
        "sheet_index": 1,
        "start_col": 61,
        "rows": list(range(4, 34)),
        "col_start": 61,
        "col_end": 64,
        "headers": ["应用名称", "下载成功", "下载失败", "下载失败率"],
    },
    "应用商店deb应用更新下载应用排行Top30及下载失败率-周累计": {
        "sheet_index": 1,
        "start_col": 65,
        "rows": list(range(4, 34)),
        "col_start": 65,
        "col_end": 68,
        "headers": ["应用名称", "下载成功", "下载失败", "下载失败率"],
    },
    "应用商店deb应用安装失败次数排行Top5及安装失败率-周累计": {
        "sheet_index": 1,
        "start_col": 69,
        "rows": list(range(4, 9)),
        "col_start": 69,
        "col_end": 71,
        "headers": ["应用名称", "安装失败次数", "安装失败率"],
    },
    "应用商店玲珑应用新增下载应用排行Top10及下载失败率-周累计": {
        "sheet_index": 1,
        "start_col": 72,
        "rows": list(range(4, 14)),
        "col_start": 72,
        "col_end": 75,
        "headers": ["应用名称", "下载成功", "下载失败", "下载失败率"],
    },
    "应用商店玲珑应用更新下载应用排行Top10及下载失败率-周累计": {
        "sheet_index": 1,
        "start_col": 76,
        "rows": list(range(4, 14)),
        "col_start": 76,
        "col_end": 79,
        "headers": ["应用名称", "下载成功", "下载失败", "下载失败率"],
    },
    "应用商店玲珑应用安装失败次数排行Top5及安装失败率-周累计": {
        "sheet_index": 1,
        "start_col": 80,
        "rows": list(range(4, 9)),
        "col_start": 80,
        "col_end": 82,
        "headers": ["应用名称", "安装失败次数", "安装失败率"],
    },
}


def canonical_block_name(name: str) -> str:
    return CANONICAL_BLOCK_NAMES.get(name, name)


def ensure_output_dir() -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    return OUTPUT_DIR


def today_stamp() -> str:
    return datetime.now().strftime("%Y%m%d")


def slugify(text: str) -> str:
    lowered = text.strip().lower()
    lowered = re.sub(r"\s+", "-", lowered)
    lowered = re.sub(r"[^0-9a-z\u4e00-\u9fff_-]+", "-", lowered)
    lowered = re.sub(r"-{2,}", "-", lowered).strip("-")
    return lowered or "output"


def discover_workbook(explicit: str = "") -> Path:
    candidates = [explicit] if explicit else []
    candidates.extend(
        [
            str(ROOT / "core_data_summary.xlsx"),
            str(DOWNLOADS_DIR / "核心数据汇总.xlsx"),
            str(DOWNLOADS_DIR / "核心数据汇总(1).xlsx"),
        ]
    )
    for candidate in candidates:
        if not candidate:
            continue
        path = Path(candidate).expanduser()
        if path.exists():
            return path
    raise FileNotFoundError("未找到可用 workbook，请显式传入 --workbook")


def write_json(path: str | Path, data: Any) -> Path:
    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return out


def read_json(path: str | Path) -> Any:
    return json.loads(Path(path).read_text(encoding="utf-8"))


def to_number(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if not text:
        return None
    if text.endswith("%"):
        return round(float(text[:-1]) / 100.0, 4)
    try:
        if re.fullmatch(r"-?\d+", text):
            return int(text)
        if re.fullmatch(r"-?\d+\.\d+", text):
            return float(text)
    except ValueError:
        return value
    return value


def normalise_scalar(value: Any) -> Any:
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return to_number(value)


def normalise_matrix(rows: list[list[Any]]) -> list[list[Any]]:
    return [[normalise_scalar(cell) for cell in row] for row in rows]


def pretty_json(value: Any) -> str:
    if value in ("", None, []):
        return ""
    return json.dumps(value, ensure_ascii=False)


@dataclass
class WorkbookSnapshot:
    path: Path

    def __post_init__(self) -> None:
        self.workbook = load_workbook(self.path, data_only=True)
        self.sheets = self.workbook.worksheets

    def block_layout(self, block_name: str) -> dict[str, Any]:
        canonical = canonical_block_name(block_name)
        if canonical not in BLOCK_LAYOUTS:
            raise KeyError(f"未登记 block: {block_name}")
        return BLOCK_LAYOUTS[canonical]

    def block_meta(self, block_name: str) -> dict[str, str]:
        canonical = canonical_block_name(block_name)
        layout = self.block_layout(canonical)
        ws = self.sheets[layout["sheet_index"]]
        raw = ws.cell(1, layout["start_col"]).value or canonical
        lines = [line.strip() for line in str(raw).splitlines() if line.strip()]
        query_path = ""
        query_rule = ""
        if len(lines) >= 2 and lines[1].startswith("查询报表："):
            query_path = lines[1].replace("查询报表：", "", 1)
        if len(lines) >= 3:
            parts = []
            for line in lines[2:]:
                if line.startswith("查询条件："):
                    parts.append(line.replace("查询条件：", "", 1))
                else:
                    parts.append(line)
            query_rule = "；".join(parts)
        return {
            "block_name": canonical,
            "sheet_name": ws.title,
            "query_path": query_path,
            "query_rule": query_rule,
        }

    def block_headers(self, block_name: str) -> list[str]:
        canonical = canonical_block_name(block_name)
        layout = self.block_layout(canonical)
        return list(layout["headers"])

    def block_matrix(self, block_name: str) -> list[list[Any]]:
        canonical = canonical_block_name(block_name)
        layout = self.block_layout(canonical)
        ws = self.sheets[layout["sheet_index"]]
        rows: list[list[Any]] = []
        for row_no in layout["rows"]:
            rows.append([ws.cell(row_no, col).value for col in range(layout["col_start"], layout["col_end"] + 1)])
        return rows


def build_compare_item(
    snapshot: WorkbookSnapshot,
    block_name: str,
    *,
    status: str,
    reason: str,
    key_dims: str = "",
    page_value: Any = "",
    request_value: Any = "",
    export_value: Any = "",
    workbook_value: Any | None = None,
) -> dict[str, Any]:
    canonical = canonical_block_name(block_name)
    meta = snapshot.block_meta(canonical)
    workbook_matrix = normalise_matrix(snapshot.block_matrix(canonical))
    return {
        "block_name": canonical,
        "sheet_name": meta["sheet_name"],
        "query_path": meta["query_path"],
        "query_rule": meta["query_rule"],
        "key_dims": key_dims,
        "workbook_value": workbook_matrix if workbook_value is None else workbook_value,
        "page_value": page_value,
        "request_value": request_value,
        "export_value": export_value,
        "status": status,
        "reason": reason,
    }
