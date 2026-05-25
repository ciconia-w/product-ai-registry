#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from core_data_common import BLOCK_LAYOUTS, OUTPUT_DIR, canonical_block_name, discover_workbook, pretty_json


def _write_matrix(ws, start_row: int, start_col: int, matrix: list[list[Any]]) -> None:
    for row_offset, row in enumerate(matrix):
        for col_offset, value in enumerate(row):
            ws.cell(start_row + row_offset, start_col + col_offset).value = value


def _matrix_shape(matrix: Any) -> tuple[int, int]:
    if not isinstance(matrix, list) or not matrix:
        return (0, 0)
    if not isinstance(matrix[0], list):
        return (0, 0)
    return (len(matrix), len(matrix[0]))


def _pick_workbook_style_matrix(layout: dict[str, Any], item: dict[str, Any]) -> list[list[Any]]:
    expected = (len(layout["rows"]), layout["col_end"] - layout["col_start"] + 1)
    for key in ("request_value", "page_value", "export_value", "workbook_value"):
        matrix = item.get(key)
        if _matrix_shape(matrix) == expected:
            return matrix
    workbook_value = item.get("workbook_value")
    if _matrix_shape(workbook_value)[0] >= expected[0] and _matrix_shape(workbook_value)[1] >= expected[1]:
        return [row[: expected[1]] for row in workbook_value[: expected[0]]]
    return []


def write_summary_sheet(wb: Workbook, items: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("核对汇总")
    headers = [
        "block_name",
        "query_path",
        "query_rule",
        "key_dims",
        "workbook_value",
        "page_value",
        "request_value",
        "export_value",
        "status",
        "reason",
    ]
    ws.append(headers)
    for item in items:
        ws.append([
            item.get("block_name", ""),
            item.get("query_path", ""),
            item.get("query_rule", ""),
            item.get("key_dims", ""),
            pretty_json(item.get("workbook_value", "")),
            pretty_json(item.get("page_value", "")),
            pretty_json(item.get("request_value", "")),
            pretty_json(item.get("export_value", "")),
            item.get("status", ""),
            item.get("reason", ""),
        ])


def write_workbook_style_sheet(output_path: Path, workbook_path: Path, items: list[dict[str, Any]]) -> None:
    source = load_workbook(workbook_path)
    block_map = {canonical_block_name(item["block_name"]): item for item in items}
    for block_name, layout in BLOCK_LAYOUTS.items():
        item = block_map.get(block_name)
        if not item:
            continue
        matrix = _pick_workbook_style_matrix(layout, item)
        if not matrix:
            continue
        ws = source.worksheets[layout["sheet_index"]]
        _write_matrix(ws, layout["rows"][0], layout["col_start"], matrix)

    if "核对汇总" in source.sheetnames:
        del source["核对汇总"]
    write_summary_sheet(source, items)
    source.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--compare-json", default=str(OUTPUT_DIR / "core_data_compare.json"))
    parser.add_argument("--workbook", default="")
    parser.add_argument("--output-xlsx", default=str(OUTPUT_DIR / "core_data_final_table.xlsx"))
    args = parser.parse_args()

    items = json.loads(Path(args.compare_json).read_text(encoding="utf-8"))
    workbook_path = discover_workbook(args.workbook)
    out = Path(args.output_xlsx)
    out.parent.mkdir(parents=True, exist_ok=True)
    write_workbook_style_sheet(out, workbook_path, items)
    print(out)


if __name__ == "__main__":
    main()
